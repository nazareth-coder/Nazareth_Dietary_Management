# exporter.py
from datetime import datetime
from typing import List, Tuple, Any

from openpyxl import Workbook

from constants import COLUMNS
from db_utils import load_patients
import excel_utils

MONTH_NAMES = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']


def _append_row(ws, row: Tuple[Any, ...]):
    ws.append(list(row))


def _parse_month(date_str: str) -> int | None:
    if not date_str:
        return None
    try:
        # Expecting YYYY-MM-DD
        dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
        return dt.month
    except Exception:
        return None


def export_db_to_excel(export_path: str):
    """
    Export all patients from DB into a fresh workbook with:
    - Main
    - Monthly sheets JAN..DEC
    - Half-year sheets JAN-JUNE, JULY-DEC
    Then run the existing summary generator to produce the metrics table.
    """
    # Load DB data as tuples in COLUMNS order
    rows: List[Tuple[Any, ...]] = load_patients()

    wb = Workbook()
    # Prepare sheets and headers
    ws_main = wb.active
    ws_main.title = 'Main'
    ws_main.append(COLUMNS)

    for m in MONTH_NAMES:
        ws = wb.create_sheet(m)
        ws.append(COLUMNS)

    ws_h1 = wb.create_sheet('JAN-JUNE')
    ws_h1.append(COLUMNS)
    ws_h2 = wb.create_sheet('JULY-DEC')
    ws_h2.append(COLUMNS)

    # Distribute rows
    date_idx = COLUMNS.index('Date of Visit') if 'Date of Visit' in COLUMNS else None
    for row in rows:
        _append_row(ws_main, row)
        month = None
        if date_idx is not None and date_idx < len(row):
            month = _parse_month(str(row[date_idx]) if row[date_idx] is not None else '')
        if month and 1 <= month <= 12:
            # Monthly
            ws_month = wb[MONTH_NAMES[month - 1]]
            _append_row(ws_month, row)
            # Half-year
            if 1 <= month <= 6:
                _append_row(ws_h1, row)
            else:
                _append_row(ws_h2, row)

    # Save the base workbook first
    wb.save(export_path)

    # Run existing summary function on the generated file by temporarily
    # pointing excel_utils.EXCEL_FILE to export_path
    try:
        original = getattr(excel_utils, 'EXCEL_FILE', None)
        # excel_utils imported EXCEL_FILE at module level from constants, so set the module var
        excel_utils.EXCEL_FILE = export_path
        excel_utils.add_summary_table_to_all_sheets()
    finally:
        if original is not None:
            excel_utils.EXCEL_FILE = original
