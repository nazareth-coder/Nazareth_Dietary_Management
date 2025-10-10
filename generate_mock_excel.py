import openpyxl
from openpyxl import Workbook
from random import randint, choice
from datetime import datetime, timedelta

from constants import COLUMNS

names = ['John Smith', 'Jane Doe', 'Alice Lee', 'Bob Cruz', 'Carla Reyes', 'Mike Tan', 'Anna Lim', 'Chris Yu', 'Nina Gomez', 'Paul Chan', 'Grace Ong', 'Leo Santos', 'Mia Cruz', 'Ray Lee', 'Ella Tan', 'Sam Yu', 'Lily Ong', 'Ivan Lim', 'Rita Chan', 'Vince Gomez']
diagnoses = ['Pneumonia', 'Dengue', 'Anemia', 'Diabetes', 'Malnutrition', 'Asthma', 'UTI', 'Appendicitis', 'Sepsis', 'Hypertension']
wards = ['ICU', 'NICU', 'Observation', 'Recovery', 'Caring 1', 'Caring 2', 'Caring 3', 'Caring 4']
subspecialties = ['Pulmo', 'Infectious', 'Hema/Oncology', 'Endo/Metabolic Disorder', 'General Service', 'Surgery', 'Cardio', 'Neuro', 'Gastro', 'Renal']
age_groups = ['0-4 years old', '5-9 years old', '10-14 years old', '15-18 years old', '19-29 years old', '30-39 years old', '40-49 years old', '50-59 years old', '60 years old and above']
types_of_visit = ['Routine', 'Referred By Doctor', 'Referred By Nurse Through Screening form']
purposes = ['Monitoring', 'Reassessment', 'Initial Assessment', 'Nutrition Education']
nutrition_support = ['Yes', 'No']
nutritional_status = ['Normal', 'Underweight', 'Stunting', 'Overweight', 'Obese', 'MAM', 'SAM']
bowel_movements = ['Normal BM', 'Loose BM of >3x in 24h', 'No BM for 48hrs']
emesis = ['Present', 'Absent']
abdominal_distention = ['Present', 'Absent']
biochem = ['Normal Lab Values', 'Elevated Liver Enzymes', 'Hypoalbuminemia', 'Elevated Serum Glucose', 'Hypothyroidism', 'Electrolyte Derangement']
rnd_mgmt = ['Maintain Current Feeding', 'Change of Feeding Prescription', 'Provision of Modulars']
encoded_by = ['RB', 'LA', 'JD', 'DM', 'Inter 1', 'Inter 2', 'Intern 3', 'Intern 4']
WITH_DOCUMENTS_OPTIONS = ['Yes', 'No']
GIVEN_NCP_OPTIONS = ['Yes', 'No']

mock_data = []
base_date = datetime(2025, 1, 1)
diet_prescriptions_options = ['Low sodium', 'High protein', 'NPO', 'Soft diet', 'Diabetic', 'Regular', '']
for i in range(1, 499):
    name = choice(names) + f" {i}"
    sex = choice(['male', 'female'])
    age = randint(1, 75)
    if age < 10:
        age_group = '0-4 years old'
    elif age < 10:
        age_group = '5-9 years old'
    elif age < 15:
        age_group = '10-14 years old'
    elif age < 19:
        age_group = '15-18 years old'
    elif age < 30:
        age_group = '19-29 years old'
    elif age < 40:
        age_group = '30-39 years old'
    elif age < 50:
        age_group = '40-49 years old'
    elif age < 60:
        age_group = '50-59 years old'
    else:
        age_group = '60 years old and above'
    # Distribute admission dates evenly across all months
    month = ((i-1) % 12) + 1
    day = randint(1, 28)  # Safe for all months
    ward_adm_dt = datetime(2025, month, day)
    visit_dt = ward_adm_dt + timedelta(days=randint(0, 10))
    diagnosis = choice(diagnoses)
    wsupport = choice(nutrition_support)
    ward = choice(wards)
    subspec = choice(subspecialties)
    height = randint(80, 180)
    weight = randint(10, 90)
    visit_type = choice(types_of_visit)
    purpose = choice(purposes)
    zscore = round(randint(-30, 30)/10, 1)
    bmi_pct = randint(1, 99)
    nut_status = choice(nutritional_status)
    bm = choice(bowel_movements)
    em = choice(emesis)
    abd = choice(abdominal_distention)
    bio = choice(biochem)
    rnd = choice(rnd_mgmt)
    diet_presc = choice(diet_prescriptions_options)
    with_docs = choice(WITH_DOCUMENTS_OPTIONS)
    given_ncp = choice(GIVEN_NCP_OPTIONS)
    encoder = choice(encoded_by)
    encoded_dt = visit_dt.strftime('%Y-%m-%d')
    last_updated = datetime(2025, 7, 1, 10, 0, 0).strftime('%Y-%m-%d %H:%M:%S')
    row = [i, name, sex, age, age_group, ward_adm_dt.strftime('%Y-%m-%d'), visit_dt.strftime('%Y-%m-%d'), diagnosis, wsupport, ward, subspec, height, weight, visit_type, purpose, zscore, bmi_pct, nut_status, bm, em, abd, bio, rnd, diet_presc, with_docs, given_ncp, encoder, encoded_dt, last_updated]
    mock_data.append(row)

wb = Workbook()
ws_main = wb.active
ws_main.title = 'Main'
ws_main.append(COLUMNS)
for row in mock_data:
    ws_main.append(row)

ws_jan_june = wb.create_sheet('JAN-JUNE')
ws_jan_june.append(COLUMNS)
ws_july_dec = wb.create_sheet('JULY-DEC')
ws_july_dec.append(COLUMNS)

# Create monthly sheets
month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
ws_months = {}
for m in month_names:
    ws_months[m] = wb.create_sheet(m)
    ws_months[m].append(COLUMNS)

for row in mock_data:
    # Half years
    try:
        month = int(str(row[5]).split('-')[1])
    except Exception:
        month = 0
    if 1 <= month <= 6:
        ws_jan_june.append(row)
    if 7 <= month <= 12:
        ws_july_dec.append(row)
    # Months
    if 1 <= month <= 12:
        ws_months[month_names[month-1]].append(row)

wb.save('Dietary_Report_MOCK.xlsx')
