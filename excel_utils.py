# excel_utils.py
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from constants import COLUMNS, EXCEL_FILE
from openpyxl.styles import Protection

def init_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Main'
        ws.append(COLUMNS)
        # Create monthly sheets
        month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        for m in month_names:
            wb.create_sheet(m)
            wb[m].append(COLUMNS)
        # (Optional) Keep half-year sheets
        wb.create_sheet('JAN-JUNE')
        wb['JAN-JUNE'].append(COLUMNS)
        wb.create_sheet('JULY-DEC')
        wb['JULY-DEC'].append(COLUMNS)
        wb.save(EXCEL_FILE)

def load_patients():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Main']
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Only load if at least one main data column is not None or blank
        if any((cell is not None and str(cell).strip() != '') for cell in row[:len(COLUMNS)]):
            data.append(row)
    return data

class ExcelFileOpenError(Exception):
    pass

def add_patient(item):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Main']
    ws.append(item)
    # Sync to half-year sheet
    try:
        date_of_visit = item[COLUMNS.index('Date of Visit')]
        month = int(str(date_of_visit).split('-')[1])
    except Exception:
        month = 0
    if 1 <= month <= 6 and 'JAN-JUNE' in wb.sheetnames:
        wb['JAN-JUNE'].append(item)
    elif 7 <= month <= 12 and 'JULY-DEC' in wb.sheetnames:
        wb['JULY-DEC'].append(item)
    # Sync to monthly sheet
    month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    if 1 <= month <= 12:
        sheet_name = month_names[month-1]
        if sheet_name in wb.sheetnames:
            wb[sheet_name].append(item)
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise ExcelFileOpenError("The Excel database file is currently open in another program. Please close it and try again.")

def update_patient(item_id, updated_item):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Main']
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(item_id):
            for idx, val in enumerate(updated_item):
                row[idx].value = val
            break
    # Remove from both half-year sheets and add to correct one
    for sheet_name in ['JAN-JUNE', 'JULY-DEC']:
        if sheet_name in wb.sheetnames:
            ws_half = wb[sheet_name]
            for idx, row in enumerate(ws_half.iter_rows(min_row=2), start=2):
                if str(row[0].value) == str(item_id):
                    ws_half.delete_rows(idx)
                    break
    # Add updated row to correct half-year sheet
    try:
        date_of_visit = updated_item[COLUMNS.index('Date of Visit')]
        month = int(str(date_of_visit).split('-')[1])
    except Exception:
        month = 0
    if 1 <= month <= 6 and 'JAN-JUNE' in wb.sheetnames:
        wb['JAN-JUNE'].append(updated_item)
    elif 7 <= month <= 12 and 'JULY-DEC' in wb.sheetnames:
        wb['JULY-DEC'].append(updated_item)
    # Remove from monthly sheets and add to the correct one
    month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    for m in month_names:
        if m in wb.sheetnames:
            ws_month = wb[m]
            for idx, row in enumerate(ws_month.iter_rows(min_row=2), start=2):
                if str(row[0].value) == str(item_id):
                    ws_month.delete_rows(idx)
                    break
    if 1 <= month <= 12:
        sheet_name = month_names[month-1]
        if sheet_name in wb.sheetnames:
            wb[sheet_name].append(updated_item)
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise ExcelFileOpenError("The Excel database file is currently open in another program. Please close it and try again.")

def delete_patient(item_id):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Main']
    admission_month = None
    # Find the admission month before deleting from Main
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(item_id):
            try:
                date_of_visit = row[COLUMNS.index('Date of Visit')].value
                if date_of_visit:
                    admission_month = int(str(date_of_visit).split('-')[1])
            except Exception:
                admission_month = None
            ws.delete_rows(idx)
            break
    # Remove from half-year sheets
    for sheet_name in ['JAN-JUNE', 'JULY-DEC']:
        if sheet_name in wb.sheetnames:
            ws_half = wb[sheet_name]
            for idx, row in enumerate(ws_half.iter_rows(min_row=2), start=2):
                if str(row[0].value) == str(item_id):
                    ws_half.delete_rows(idx)
                    break
    # Remove from monthly sheet
    month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    if admission_month and 1 <= admission_month <= 12:
        sheet_name = month_names[admission_month-1]
        if sheet_name in wb.sheetnames:
            ws_month = wb[sheet_name]
            for idx, row in enumerate(ws_month.iter_rows(min_row=2), start=2):
                if str(row[0].value) == str(item_id):
                    ws_month.delete_rows(idx)
                    break
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise ExcelFileOpenError("The Excel database file is currently open in another program. Please close it and try again.")

def export_excel(export_path):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    wb.save(export_path)


def add_summary_table_to_all_sheets():
    import openpyxl.utils
    wb = openpyxl.load_workbook(EXCEL_FILE)
    metrics = [
        'Number of patients admitted',
        'Number of nutritionally-at-risk (NAR) patients',
        '  a. Wasting',
        '    i. Moderate acute malnutrition',
        '    ii. Severe acute malnutrition',
        '  b. Stunting',
        '  c. Underweight',
        '  d. Overweight',
        '  e. Obese',
        '  f. Disease and other co-morbidities (Please specify)',
        '    i.',
        '    ii.',
        '    iii.',
        '    iv.',
        'Number of NAR patients given nutrition screening (by the nurse)',
        'Number of patients given nutrition assessment',
        'Number of patients given nutrition intervention',
        'Number of patients with nutrition documentation',
        'Number of patients given nutrition care process (ADIME)',
    ]
    age_groups = ['0-4 years old', '5-9 years old', '10-14 years old', '15-18 years old', '19-29 years old', '30-39 years old', '40-49 years old', '50-59 years old', '60 years old and above']
    table_age_groups = ['0-4', '5-9', '10-14', '15-18', '19-29', '30-39', '40-49', '50-59', '60 & above']
    sex = ['male', 'female']
    sex_table = ['M', 'F']
    month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    sheet_names = ['Main', 'JAN-JUNE', 'JULY-DEC'] + month_names

    # Map metric index to column in data
    metric_to_col = {
        14: 'Type Of Visit',           # 'Number of NAR patients given nutrition screening (by the nurse)'
        15: 'Purpose of Visit',        # 'Number of patients given nutrition assessment'
        16: 'RND Dietary Management',  # 'Number of patients given nutrition intervention'
    }

    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            start_col = 31  # AE
            start_row = 1

            # --- Write header row ---
            ws.cell(row=start_row, column=start_col, value='Metrics')
            col = start_col + 1
            for group in table_age_groups:
                for s in sex_table:
                    ws.cell(row=start_row, column=col, value=f'{group}\n{s}')
                    col += 1
            ws.cell(row=start_row, column=col, value='Subtotal\nM')
            ws.cell(row=start_row, column=col+1, value='Subtotal\nF')
            ws.cell(row=start_row, column=col+2, value='TOTAL')

            # --- Prepare data structures for counts ---
            counts = {}
            for metric_idx in metric_to_col:
                counts[metric_idx] = [[0, 0] for _ in range(len(table_age_groups))]  # [age_group][sex], M=0, F=1

            # --- Read data and count ---
            # Find column indices in sheet
            header_row = [cell.value for cell in ws[1]]
            col_indices = {col: header_row.index(col) for col in metric_to_col.values() if col in header_row}
            age_idx = header_row.index('Age Group') if 'Age Group' in header_row else None
            sex_idx = header_row.index('Sex') if 'Sex' in header_row else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                # Map age group to index
                age_val = row[age_idx] if age_idx is not None else ''
                sex_val = str(row[sex_idx]).strip().lower() if sex_idx is not None and row[sex_idx] else ''
                try:
                    age_idx_tbl = age_groups.index(age_val)
                except Exception:
                    continue
                sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                if sex_tbl_idx is None:
                    continue
                # For each metric, check if value is non-empty
                for metric_idx, col_name in metric_to_col.items():
                    col_idx = col_indices.get(col_name)
                    if col_idx is not None and row[col_idx]:
                        counts[metric_idx][age_idx_tbl][sex_tbl_idx] += 1

            # --- Write metrics and counts to table ---
            for i, metric in enumerate(metrics):
                ws.cell(row=start_row + 1 + i, column=start_col, value=metric)
                if i in metric_to_col:
                    # Fill counts for this metric
                    subtotal_m = 0
                    subtotal_f = 0
                    total = 0
                    for ag in range(len(table_age_groups)):
                        for sx in range(2):
                            val = counts[i][ag][sx]
                            ws.cell(row=start_row + 1 + i, column=start_col + 1 + ag*2 + sx, value=val)
                            if sx == 0:
                                subtotal_m += val
                            else:
                                subtotal_f += val
                            total += val
                    # Write subtotals and total
                    ws.cell(row=start_row + 1 + i, column=start_col + 1 + len(table_age_groups)*2, value=subtotal_m)
                    ws.cell(row=start_row + 1 + i, column=start_col + 2 + len(table_age_groups)*2, value=subtotal_f)
                    ws.cell(row=start_row + 1 + i, column=start_col + 3 + len(table_age_groups)*2, value=total)
                else:
                    # Fill empty cells for non-implemented metrics
                    for c in range(start_col + 1, start_col + 1 + len(table_age_groups)*2 + 3):
                        ws.cell(row=start_row + 1 + i, column=c, value=None)


            # --- Fill 'Number of patients admitted' row ---
            admitted_row_idx = 1  # Row index in summary table for 'Number of patients admitted' (first metric)
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Find column indices for Age Group and Sex in the patient data table
                    header_row = [cell.value for cell in ws[1]]
                    try:
                        age_col_idx = header_row.index('Age Group')
                        sex_col_idx = header_row.index('Sex')
                    except ValueError:
                        continue  # Skip if columns not found
                    # Set up counters
                    admitted_counts = [[0, 0] for _ in age_groups]  # [age_group][sex: 0=M, 1=F]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        age_val = row[age_col_idx]
                        sex_val = str(row[sex_col_idx]).strip().lower() if row[sex_col_idx] else ''
                        # Map to table index
                        try:
                            age_tbl_idx = age_groups.index(age_val)
                        except Exception:
                            continue
                        sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                        if sex_tbl_idx is None:
                                continue
                        admitted_counts[age_tbl_idx][sex_tbl_idx] += 1
                        # Write counts to summary table
                        start_col = 31  # AE
                        start_row = 1
                        col_ptr = start_col + 1
                        subtotal_m = 0
                        subtotal_f = 0
                        total = 0
                        for ag in range(len(age_groups)):
                            for sx in range(2):
                                val = admitted_counts[ag][sx]
                                ws.cell(row=start_row + admitted_row_idx, column=col_ptr, value=val)
                                if sx == 0:
                                    subtotal_m += val
                                else:
                                    subtotal_f += val
                                total += val
                                col_ptr += 1
                        # Subtotals and total
                        ws.cell(row=start_row + admitted_row_idx, column=col_ptr, value=subtotal_m)
                        ws.cell(row=start_row + admitted_row_idx, column=col_ptr+1, value=subtotal_f)
                        ws.cell(row=start_row + admitted_row_idx, column=col_ptr+2, value=total)

            # --- Fill Stunting, Underweight, Overweight, Obese rows ---
            metric_names = {
                "b. Stunting": "Stunting",
                "c. Underweight": "Underweight",
                "d. Overweight": "Overweight",
                "e. Obese": "Obese"
            }
            # Find the row index in the summary table for each metric
            metric_row_indices = {}
            for i, metric in enumerate(metrics):
                for k, v in metric_names.items():
                    if metric.strip().lower().startswith(k.lower()):
                        metric_row_indices[v] = i + 1  # +1 because header is at row 1

            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Find column indices for Age Group, Sex, and Nutritional Status
                    header_row = [cell.value for cell in ws[1]]
                    try:
                        age_col_idx = header_row.index('Age Group')
                        sex_col_idx = header_row.index('Sex')
                        status_col_idx = header_row.index('Nutritional Status')
                    except ValueError:
                        continue  # Skip if columns not found
                    # Set up counters for each metric
                    metric_counts = {k: [[0, 0] for _ in age_groups] for k in metric_names.values()}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        age_val = row[age_col_idx]
                        sex_val = str(row[sex_col_idx]).strip().lower() if row[sex_col_idx] else ''
                        status_val = str(row[status_col_idx]).strip().capitalize() if row[status_col_idx] else ''
                        # Map to table index
                        try:
                            age_tbl_idx = age_groups.index(age_val)
                        except Exception:
                            continue
                        sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                        if sex_tbl_idx is None:
                            continue
                        for metric_key, metric_label in metric_names.items():
                            if status_val.lower() == metric_label.lower():
                                metric_counts[metric_label][age_tbl_idx][sex_tbl_idx] += 1
                    # Write counts to summary table
                    start_col = 31  # AE
                    start_row = 1
                    for metric_label, row_idx in metric_row_indices.items():
                        col_ptr = start_col + 1
                        subtotal_m = 0
                        subtotal_f = 0
                        total = 0
                        for ag in range(len(age_groups)):
                            for sx in range(2):
                                val = metric_counts[metric_label][ag][sx]
                                ws.cell(row=start_row + row_idx, column=col_ptr, value=val)
                                if sx == 0:
                                    subtotal_m += val
                                else:
                                    subtotal_f += val
                                total += val
                                col_ptr += 1
                        ws.cell(row=start_row + row_idx, column=col_ptr, value=subtotal_m)
                        ws.cell(row=start_row + row_idx, column=col_ptr+1, value=subtotal_f)
                        ws.cell(row=start_row + row_idx, column=col_ptr+2, value=total)

                      # --- Fill "Number of nutritionally-at-risk (NAR) patients" row ---
    # Find the row index in metrics for this label
    nar_metric_row = None
    for i, metric in enumerate(metrics):
        if metric.strip().lower().startswith("number of nutritionally-at-risk"):
            nar_metric_row = i + 1  # +1 for header
            break

    if nar_metric_row:
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row = [cell.value for cell in ws[1]]
                try:
                    age_col_idx = header_row.index('Age Group')
                    sex_col_idx = header_row.index('Sex')
                    status_col_idx = header_row.index('Nutritional Status')
                except ValueError:
                    continue  # Skip if columns not found
                nar_counts = [[0, 0] for _ in age_groups]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    age_val = row[age_col_idx]
                    sex_val = str(row[sex_col_idx]).strip().lower() if row[sex_col_idx] else ''
                    status_val = str(row[status_col_idx]).strip().lower() if row[status_col_idx] else ''
                    # Only count if not normal
                    if status_val == "normal":
                        continue
                    try:
                        age_tbl_idx = age_groups.index(age_val)
                    except Exception:
                        continue
                    sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                    if sex_tbl_idx is None:
                        continue
                    nar_counts[age_tbl_idx][sex_tbl_idx] += 1
                # Write counts to summary table
                start_col = 31  # AE
                start_row = 1
                col_ptr = start_col + 1
                subtotal_m = 0
                subtotal_f = 0
                total = 0
                for ag in range(len(age_groups)):
                    for sx in range(2):
                        val = nar_counts[ag][sx]
                        ws.cell(row=start_row + nar_metric_row, column=col_ptr, value=val)
                        if sx == 0:
                            subtotal_m += val
                        else:
                            subtotal_f += val
                        total += val
                        col_ptr += 1
                ws.cell(row=start_row + nar_metric_row, column=col_ptr, value=subtotal_m)
                ws.cell(row=start_row + nar_metric_row, column=col_ptr+1, value=subtotal_f)
                ws.cell(row=start_row + nar_metric_row, column=col_ptr+2, value=total)



                # --- Fill Moderate and Severe acute malnutrition rows ---
    mam_sam_metrics = {
        "i. Moderate acute malnutrition": "MAM",
        "ii. Severe acute malnutrition": "SAM"
    }
    # Find row indices for these metrics
    mam_sam_row_indices = {}
    for i, metric in enumerate(metrics):
        for label in mam_sam_metrics:
            if metric.strip().lower().startswith(label.lower()):
                mam_sam_row_indices[label] = i + 1  # +1 for header

    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = [cell.value for cell in ws[1]]
            try:
                age_col_idx = header_row.index('Age Group')
                sex_col_idx = header_row.index('Sex')
                status_col_idx = header_row.index('Nutritional Status')
            except ValueError:
                continue  # Skip if columns not found
            # Set up counters
            mam_counts = [[0, 0] for _ in age_groups]
            sam_counts = [[0, 0] for _ in age_groups]
            for row in ws.iter_rows(min_row=2, values_only=True):
                age_val = row[age_col_idx]
                sex_val = str(row[sex_col_idx]).strip().lower() if row[sex_col_idx] else ''
                status_val = str(row[status_col_idx]).strip().upper() if row[status_col_idx] else ''
                try:
                    age_tbl_idx = age_groups.index(age_val)
                except Exception:
                    continue
                sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                if sex_tbl_idx is None:
                    continue
                if status_val == "MAM":
                    mam_counts[age_tbl_idx][sex_tbl_idx] += 1
                elif status_val == "SAM":
                    sam_counts[age_tbl_idx][sex_tbl_idx] += 1
            # Write counts to summary table
            start_col = 31  # AE
            start_row = 1
            # Moderate acute malnutrition
            if "i. Moderate acute malnutrition" in mam_sam_row_indices:
                row_idx = mam_sam_row_indices["i. Moderate acute malnutrition"]
                col_ptr = start_col + 1
                subtotal_m = subtotal_f = total = 0
                for ag in range(len(age_groups)):
                    for sx in range(2):
                        val = mam_counts[ag][sx]
                        ws.cell(row=start_row + row_idx, column=col_ptr, value=val)
                        if sx == 0:
                            subtotal_m += val
                        else:
                            subtotal_f += val
                        total += val
                        col_ptr += 1
                ws.cell(row=start_row + row_idx, column=col_ptr, value=subtotal_m)
                ws.cell(row=start_row + row_idx, column=col_ptr+1, value=subtotal_f)
                ws.cell(row=start_row + row_idx, column=col_ptr+2, value=total)
            # Severe acute malnutrition
            if "ii. Severe acute malnutrition" in mam_sam_row_indices:
                row_idx = mam_sam_row_indices["ii. Severe acute malnutrition"]
                col_ptr = start_col + 1
                subtotal_m = subtotal_f = total = 0
                for ag in range(len(age_groups)):
                    for sx in range(2):
                        val = sam_counts[ag][sx]
                        ws.cell(row=start_row + row_idx, column=col_ptr, value=val)
                        if sx == 0:
                            subtotal_m += val
                        else:
                            subtotal_f += val
                        total += val
                        col_ptr += 1
                ws.cell(row=start_row + row_idx, column=col_ptr, value=subtotal_m)
                ws.cell(row=start_row + row_idx, column=col_ptr+1, value=subtotal_f)
                ws.cell(row=start_row + row_idx, column=col_ptr+2, value=total)

                 # --- Fill documentation and NCP metrics ---
    doc_metric = "Number of patients with nutrition documentation"
    ncp_metric = "Number of patients given nutrition care process (ADIME)"

    # Find row indices for these metrics
    doc_row_idx = ncp_row_idx = None
    for i, metric in enumerate(metrics):
        if metric.strip().lower().startswith(doc_metric.lower()):
            doc_row_idx = i + 1  # +1 for header
        if metric.strip().lower().startswith(ncp_metric.lower()):
            ncp_row_idx = i + 1

    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = [cell.value for cell in ws[1]]
            try:
                age_col_idx = header_row.index('Age Group')
                sex_col_idx = header_row.index('Sex')
                doc_col_idx = header_row.index('With Documents')
                ncp_col_idx = header_row.index('Given NCP')
            except ValueError:
                continue  # Skip if columns not found

            # Documentation counts
            doc_counts = [[0, 0] for _ in age_groups]
            ncp_counts = [[0, 0] for _ in age_groups]

            for row in ws.iter_rows(min_row=2, values_only=True):
                age_val = row[age_col_idx]
                sex_val = str(row[sex_col_idx]).strip().lower() if row[sex_col_idx] else ''
                doc_val = str(row[doc_col_idx]).strip().lower() if row[doc_col_idx] else ''
                ncp_val = str(row[ncp_col_idx]).strip().lower() if row[ncp_col_idx] else ''
                try:
                    age_tbl_idx = age_groups.index(age_val)
                except Exception:
                    continue
                sex_tbl_idx = 0 if sex_val == 'male' else 1 if sex_val == 'female' else None
                if sex_tbl_idx is None:
                    continue
                if doc_val == "yes":
                    doc_counts[age_tbl_idx][sex_tbl_idx] += 1
                if ncp_val == "yes":
                    ncp_counts[age_tbl_idx][sex_tbl_idx] += 1

            start_col = 31  # AE
            start_row = 1

            # Write documentation counts
            if doc_row_idx is not None:
                col_ptr = start_col + 1
                subtotal_m = subtotal_f = total = 0
                for ag in range(len(age_groups)):
                    for sx in range(2):
                        val = doc_counts[ag][sx]
                        ws.cell(row=start_row + doc_row_idx, column=col_ptr, value=val)
                        if sx == 0:
                            subtotal_m += val
                        else:
                            subtotal_f += val
                        total += val
                        col_ptr += 1
                ws.cell(row=start_row + doc_row_idx, column=col_ptr, value=subtotal_m)
                ws.cell(row=start_row + doc_row_idx, column=col_ptr+1, value=subtotal_f)
                ws.cell(row=start_row + doc_row_idx, column=col_ptr+2, value=total)

            # Write NCP counts
            if ncp_row_idx is not None:
                col_ptr = start_col + 1
                subtotal_m = subtotal_f = total = 0
                for ag in range(len(age_groups)):
                    for sx in range(2):
                        val = ncp_counts[ag][sx]
                        ws.cell(row=start_row + ncp_row_idx, column=col_ptr, value=val)
                        if sx == 0:
                            subtotal_m += val
                        else:
                            subtotal_f += val
                        total += val
                        col_ptr += 1
                ws.cell(row=start_row + ncp_row_idx, column=col_ptr, value=subtotal_m)
                ws.cell(row=start_row + ncp_row_idx, column=col_ptr+1, value=subtotal_f)
                ws.cell(row=start_row + ncp_row_idx, column=col_ptr+2, value=total)

    # # --- Lock summary table cells and protect the sheet ---
    # for sheet_name in sheet_names:
    #     if sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
    #         # Unlock all cells first
    #         for row in ws.iter_rows():
    #             for cell in row:
    #                 cell.protection = Protection(locked=False)

    #         num_metrics = len(metrics)
    #         num_summary_cols = len(age_groups) * 2 + 3  # 2 sexes per group + subtotals + total

    #         # Determine which rows to skip locking
    #         skip_labels = [
    #             "a. Wasting",
    #         "i. Moderate acute malnutrition",
    #         "ii. Severe acute malnutrition",
    #         "f. Disease and other co-morbidities (Please specify)",
    #         "i.",
    #         "ii.",
    #         "iii.",
    #         "iv."
    #     ]
    #     skip_row_indices = []
    #     for i, metric in enumerate(metrics):
    #         if any(metric.strip().lower().startswith(label.lower()) for label in skip_labels):
    #             skip_row_indices.append(i + 1)  # +1 for header row offset

    #     for r in range(start_row + 1, start_row + 1 + num_metrics):
    #         if (r - start_row) in skip_row_indices:
    #             continue  # Do not lock this row
    #         for c in range(start_col, start_col + 1 + num_summary_cols):
    #             ws.cell(row=r, column=c).protection = Protection(locked=True)

    #     ws.protection.sheet = True
    #     # ws.protection.password = "metricslock"
    wb.save(EXCEL_FILE)

