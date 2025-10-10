# extractor.py
# A standalone extractor/merger GUI for merging rows from a source Excel workbook
# into a target Excel workbook. Only the first len(COLUMNS) columns (data columns)
# are merged. Metrics columns are ignored. After merge, monthly and half-year
# sheets are synchronized from Main. Metrics recomputation is left to the main app.

from tkinter import Toplevel, StringVar, END
from tkinter import ttk, filedialog, messagebox
import os
import shutil
from datetime import datetime
import openpyxl
from constants import COLUMNS
import excel_utils as eu
import uuid

MONTH_NAMES = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
DATA_COLS = len(COLUMNS)

class ExtractorWindow:
    def __init__(self, parent):
        self.parent = parent
        self.win = Toplevel(parent)
        self.win.title('Data Extractor / Merger')
        self.win.geometry('640x170')
        self.win.resizable(False, False)

        self.src_path = StringVar()
        self.tgt_path = StringVar()

        # Tighter padding
        pad = {'padx': 8, 'pady': 4}

        # Grid columns: label | entry (expands) | button
        self.win.grid_columnconfigure(1, weight=1)

        ttk.Label(self.win, text='Source (.xlsx):').grid(row=0, column=0, sticky='w', **pad)
        self.src_entry = ttk.Entry(self.win, textvariable=self.src_path)
        self.src_entry.grid(row=0, column=1, sticky='we', **pad)
        ttk.Button(self.win, text='Browse...', command=self._choose_src).grid(row=0, column=2, sticky='w', **pad)

        ttk.Label(self.win, text='Target (.xlsx):').grid(row=1, column=0, sticky='w', **pad)
        self.tgt_entry = ttk.Entry(self.win, textvariable=self.tgt_path)
        self.tgt_entry.grid(row=1, column=1, sticky='we', **pad)
        ttk.Button(self.win, text='Browse...', command=self._choose_tgt).grid(row=1, column=2, sticky='w', **pad)

        # Action buttons
        ttk.Button(self.win, text='Run Merge', command=self._run_merge).grid(row=2, column=2, sticky='e', **pad)
        ttk.Button(self.win, text='Close', command=self.win.destroy).grid(row=2, column=1, sticky='e', **pad)

        # Status line
        self.status = StringVar()
        ttk.Label(self.win, textvariable=self.status, foreground='#555').grid(row=3, column=0, columnspan=3, sticky='w', padx=8)

    def _choose_src(self):
        path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
        if path:
            self.src_path.set(path)

    def _choose_tgt(self):
        path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
        if path:
            self.tgt_path.set(path)

    def _run_merge(self):
        src = self.src_path.get().strip()
        tgt = self.tgt_path.get().strip()
        if not src or not tgt:
            messagebox.showerror('Missing files', 'Please select both source and target workbooks.')
            return
        if not os.path.isfile(src) or not os.path.isfile(tgt):
            messagebox.showerror('Invalid files', 'One or both selected files do not exist.')
            return
        if os.path.abspath(src) == os.path.abspath(tgt):
            messagebox.showerror('Same file', 'Source and target must be different files.')
            return
        try:
            self.status.set('Backing up target...')
            self.win.update_idletasks()
            backup_path = backup_target(tgt)

            self.status.set('Merging rows...')
            self.win.update_idletasks()
            result = merge_workbooks(src, tgt)

            self.status.set('Syncing monthly and half-year sheets...')
            self.win.update_idletasks()
            resync_sheets(tgt)

            # Recompute metrics safely by pointing excel_utils to the target path temporarily
            self.status.set('Recomputing metrics...')
            self.win.update_idletasks()
            try:
                old_path = eu.EXCEL_FILE
                eu.EXCEL_FILE = tgt
                eu.add_summary_table_to_all_sheets()
            finally:
                try:
                    eu.EXCEL_FILE = old_path
                except Exception:
                    pass

            self.status.set('Done.')
            msg = (
                f"Merge complete.\n\n"
                f"Inserted: {result['inserted']}\n"
                f"Updated: {result['updated']}\n"
                f"Skipped: {result['skipped']}\n\n"
                f"Target backup saved to:\n{backup_path}\n\n"
                f"Metrics were recalculated based on the merged data."
            )
            messagebox.showinfo('Merge Complete', msg)
        except Exception as e:
            messagebox.showerror('Merge Error', f'An error occurred during merge:\n{e}')


def backup_target(target_path: str) -> str:
    """Create a timestamped backup copy of the target workbook in the same directory."""
    dirn = os.path.dirname(target_path)
    base = os.path.basename(target_path)
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    backup_name = f"{os.path.splitext(base)[0]}-BACKUP-{ts}.xlsx"
    backup_path = os.path.join(dirn, backup_name)
    shutil.copy2(target_path, backup_path)
    return backup_path


def merge_workbooks(src_path: str, tgt_path: str):
    """UUID-based merge of data rows from src Main into tgt Main.
    - Uses a hidden UUID stored in the blank column immediately after the data columns (col = DATA_COLS + 1).
    - Assigns UUIDs to any rows (in source or target) that are missing one.
    - Upserts by UUID. If UUID exists in target, overwrite data columns. Otherwise, append as new.
    - If inserting and Patient ID is blank, assign next available numeric ID in target.
    Returns a dict with counts and saves BOTH source and target.
    """
    wb_src = openpyxl.load_workbook(src_path)
    wb_tgt = openpyxl.load_workbook(tgt_path)

    if 'Main' not in wb_src.sheetnames or 'Main' not in wb_tgt.sheetnames:
        raise ValueError("Both source and target must have a 'Main' sheet.")

    ws_src = wb_src['Main']
    ws_tgt = wb_tgt['Main']

    uuid_col = DATA_COLS + 1  # first column after data columns

    def ensure_row_uuids(ws):
        count = 0
        for r in range(2, ws.max_row + 1):
            # if the row has any data in the first DATA_COLS cells, consider it a data row
            if any(ws.cell(row=r, column=c).value for c in range(1, DATA_COLS + 1)):
                cell = ws.cell(row=r, column=uuid_col)
                if cell.value is None or str(cell.value).strip() == '':
                    cell.value = str(uuid.uuid4())
                    count += 1
        return count

    # Assign UUIDs to any rows missing them, in both workbooks
    ensure_row_uuids(ws_tgt)
    ensure_row_uuids(ws_src)

    # Build target index by UUID and find max Patient ID
    tgt_index = {}
    max_id = 0
    for i, row in enumerate(ws_tgt.iter_rows(min_row=2), start=2):
        if not any(cell.value for cell in row[:DATA_COLS]):
            continue
        pid = row[0].value
        try:
            if pid is not None and str(pid).strip() != '':
                max_id = max(max_id, int(str(pid)))
        except Exception:
            pass
        uval = row[uuid_col - 1].value  # openpyxl rows are 0-indexed in this tuple
        if uval:
            tgt_index[str(uval)] = i

    inserted = updated = skipped = 0

    # Iterate source rows and upsert by UUID
    for r in range(2, ws_src.max_row + 1):
        data = [ws_src.cell(row=r, column=c).value for c in range(1, DATA_COLS + 1)]
        if not any(data):
            continue
        ucell = ws_src.cell(row=r, column=uuid_col)
        uval = str(ucell.value).strip() if ucell.value else ''
        if not uval:
            uval = str(uuid.uuid4())
            ucell.value = uval  # persist back to source
        if uval in tgt_index:
            tgt_row_idx = tgt_index[uval]
            # overwrite data columns EXCEPT Patient ID (col 1) to preserve target ID
            for c in range(2, DATA_COLS + 1):
                ws_tgt.cell(row=tgt_row_idx, column=c, value=data[c - 1])
            updated += 1
        else:
            # append as new
            # always assign a fresh Patient ID to avoid duplicates
            max_id += 1
            data[0] = max_id
            ws_tgt.append(data + [uval])  # include UUID cell to keep alignment
            inserted += 1

    # Prune trailing empty rows in Main
    def prune_trailing_empty(ws):
        last = ws.max_row
        while last > 1:
            # Check if all data columns are empty in this row
            if any(ws.cell(row=last, column=c).value for c in range(1, DATA_COLS + 1)):
                break
            ws.delete_rows(last)
            last -= 1

    prune_trailing_empty(ws_tgt)

    # Save both workbooks to persist UUIDs and cleaned rows
    wb_src.save(src_path)
    wb_tgt.save(tgt_path)
    return {'inserted': inserted, 'updated': updated, 'skipped': skipped}


def resync_sheets(tgt_path: str):
    """Rebuild half-year and monthly sheets' data from Main based on Ward Admission Date month.
    Does not touch metrics cells beyond the data columns.
    """
    wb = openpyxl.load_workbook(tgt_path)
    if 'Main' not in wb.sheetnames:
        wb.save(tgt_path)
        return
    ws_main = wb['Main']

    # Collect data rows
    rows = []
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        data = list(row[:DATA_COLS])
        if any(data):
            rows.append(data)

    # Ensure half-year and monthly sheets exist with headers
    def ensure_sheet(name):
        if name not in wb.sheetnames:
            wb.create_sheet(name)
            wb[name].append(COLUMNS)
        return wb[name]

    sh_first = ensure_sheet('JAN-JUNE')
    sh_second = ensure_sheet('JULY-DEC')
    month_sheets = {m: ensure_sheet(m) for m in MONTH_NAMES}

    # Clear existing data rows (keep header row)
    def clear_data(ws):
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    clear_data(sh_first)
    clear_data(sh_second)
    for m in MONTH_NAMES:
        clear_data(month_sheets[m])

    # Append rows to appropriate sheets
    for data in rows:
        try:
            wad = data[COLUMNS.index('Ward Admission Date')]
            month = int(str(wad).split('-')[1]) if wad else 0
        except Exception:
            month = 0
        if 1 <= month <= 6:
            sh_first.append(data)
        elif 7 <= month <= 12:
            sh_second.append(data)
        if 1 <= month <= 12:
            month_sheets[MONTH_NAMES[month-1]].append(data)

    wb.save(tgt_path)


def open_extractor_window(root):
    ExtractorWindow(root)
